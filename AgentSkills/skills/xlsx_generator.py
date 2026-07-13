"""
xlsx_generator.py - Excel 文件生成工具

提供从结构化数据或 Markdown 表格创建 .xlsx 文件的功能。
支持中文内容，自动美化格式（表头样式、斑马纹、列宽调整、边框）。
支持多工作表、自定义样式、自动安装依赖。
"""

import os
import re
import subprocess
import sys
from typing import Any, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════
# 样式常量（默认值）
# ══════════════════════════════════════════════════════════════

_HEADER_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")  # 浅蓝背景
_HEADER_FONT = Font(bold=True, size=11)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

_DATA_FONT = Font(size=10)
_DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center")

_THIN_BORDER = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
    bottom=Side(style="thin", color="999999"),
)

_EVEN_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")   # 浅蓝斑马纹
_ODD_ROW_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")     # 白色

_TITLE_FONT = Font(bold=True, size=14, color="2F5496")
_TITLE_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# ── 类型别名 ──
_SingleData = list[list]                            # 单工作表数据
_MultiSheetByDict = dict[str, list[list]]           # 多工作表：{sheet名: 数据}
_SheetDef = dict[str, Any]                          # 单个工作表定义
_MultiSheetByList = list[_SheetDef]                 # 多工作表：[{...}, {...}]
_SupportedData = Union[_SingleData, _MultiSheetByDict, _MultiSheetByList]


# ══════════════════════════════════════════════════════════════
# 工具函数
# ══════════════════════════════════════════════════════════════

def _auto_install_openpyxl() -> bool:
    """自动安装 openpyxl 库（如果未安装）

    检测当前环境中是否已安装 openpyxl，
    若未安装则自动通过 pip 安装。

    Returns:
        安装成功返回 True，失败返回 False
    """
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        print("📦 检测到 openpyxl 未安装，正在自动安装...")
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", "openpyxl", "-q"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            print("✅ openpyxl 安装成功！")
            # 重新导入以确保加载
            import importlib
            importlib.reload(__import__('openpyxl'))
            return True
        except Exception as e:
            print(f"❌ openpyxl 自动安装失败: {e}")
            print("💡 请手动执行: pip install openpyxl")
            return False


def _apply_cell_style(cell, font=None, fill=None, alignment=None, border=None):
    """统一设置单元格样式

    Args:
        cell: openpyxl 单元格对象
        font: 字体样式
        fill: 填充样式
        alignment: 对齐方式
        border: 边框样式
    """
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def _auto_adjust_column_widths(ws, data: list[list], title: Optional[str] = None):
    """自动调整列宽

    基于表头和数据内容的长度计算最佳列宽。
    中文字符按2个英文字符宽度计算。

    Args:
        ws: openpyxl 工作表对象
        data: 二维列表数据
        title: 可选的标题文本，影响列宽计算
    """
    if not data:
        return

    num_cols = len(data[0]) if data else 0
    for col_idx in range(1, num_cols + 1):
        max_length = 0

        # 考虑标题行长度
        if title:
            title_len = len(str(title)) // num_cols + 2
            max_length = max(max_length, title_len)

        # 遍历所有数据行
        for row in data:
            if col_idx - 1 < len(row):
                cell_text = str(row[col_idx - 1])
                # 中文字符按2个宽度计算
                char_length = 0
                for ch in cell_text:
                    if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f':
                        char_length += 2
                    else:
                        char_length += 1
                max_length = max(max_length, char_length)

        # 设置列宽（最小宽度8，最大宽度60）
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max(max_length + 3, 8), 60)


def _apply_table_formatting(
    ws,
    data: list[list],
    start_row: int = 1,
    header_font=None,
    header_fill=None,
    header_alignment=None,
    data_font=None,
    data_alignment=None,
    even_fill=None,
    odd_fill=None,
    border=None,
):
    """对工作表数据区域应用格式：表头样式、斑马纹、边框

    Args:
        ws: openpyxl 工作表对象
        data: 二维列表数据
        start_row: 数据起始行（如果前面有标题行）
        header_font: 表头字体（默认加粗）
        header_fill: 表头填充（默认浅蓝）
        header_alignment: 表头对齐（默认居中）
        data_font: 数据字体
        data_alignment: 数据对齐
        even_fill: 偶数行填充（斑马纹）
        odd_fill: 奇数行填充
        border: 单元格边框
    """
    if not data:
        return

    num_rows = len(data)
    num_cols = len(data[0]) if data else 0

    # 使用默认值或传入的自定义值
    h_font = header_font or _HEADER_FONT
    h_fill = header_fill or _HEADER_FILL
    h_align = header_alignment or _HEADER_ALIGNMENT
    d_font = data_font or _DATA_FONT
    d_align = data_alignment or _DATA_ALIGNMENT
    ev_fill = even_fill or _EVEN_ROW_FILL
    od_fill = odd_fill or _ODD_ROW_FILL
    bdr = border or _THIN_BORDER

    for row_idx, row_data in enumerate(data):
        for col_idx in range(num_cols):
            cell = ws.cell(row=start_row + row_idx, column=col_idx + 1)
            cell_value = row_data[col_idx] if col_idx < len(row_data) else ""
            cell.value = cell_value

            if row_idx == 0:
                # 表头行：加粗、浅蓝背景、居中对齐
                _apply_cell_style(
                    cell,
                    font=h_font,
                    fill=h_fill,
                    alignment=h_align,
                    border=bdr,
                )
            else:
                # 数据行：斑马纹
                row_fill = ev_fill if row_idx % 2 == 0 else od_fill
                _apply_cell_style(
                    cell,
                    font=d_font,
                    fill=row_fill,
                    alignment=d_align,
                    border=bdr,
                )


def _ensure_output_path(output_path: str) -> str:
    """确保输出路径以 .xlsx 结尾，并自动创建父目录

    Args:
        output_path: 输出路径

    Returns:
        修正后的输出路径
    """
    if not output_path.endswith(".xlsx"):
        output_path += ".xlsx"

    parent_dir = os.path.dirname(output_path)
    if parent_dir and not os.path.exists(parent_dir):
        os.makedirs(parent_dir, exist_ok=True)

    return output_path


def _format_file_size(file_path: str) -> str:
    """获取文件大小并格式化为可读字符串

    Args:
        file_path: 文件路径

    Returns:
        格式化的文件大小字符串，如 "5.2 KB"
    """
    file_size = os.path.getsize(file_path)
    if file_size < 1024:
        return f"{file_size} B"
    elif file_size < 1024 * 1024:
        return f"{file_size / 1024:.1f} KB"
    else:
        return f"{file_size / (1024 * 1024):.1f} MB"


# ══════════════════════════════════════════════════════════════
# 主函数
# ══════════════════════════════════════════════════════════════

def create_xlsx(
    data,
    output_path: str,
    sheet_name: str = "Sheet1",
    title: Optional[str] = None,
    auto_install: bool = True,
    styles: Optional[dict] = None,
) -> str:
    """创建格式美观的 Excel 文件

    🎯 **核心功能**：
    - ✅ 支持**多个工作表**（Sheet）— 传入 dict 或 list 即可
    - ✅ 支持**表头样式**（加粗、居中、浅蓝背景）
    - ✅ 支持**斑马纹**交替行颜色
    - ✅ 支持**单元格边框**和**自动列宽**
    - ✅ 支持**标题行**（可选，合并单元格）
    - ✅ 支持**自定义样式**（字体、颜色、边框等）
    - ✅ 自动**安装 openpyxl**（如果未安装）

    📦 **数据格式说明**：

    **格式一：单工作表（向后兼容）**
    ```python
    data = [["姓名", "年龄"], ["张三", 28], ["李四", 35]]
    create_xlsx(data, "/tmp/人员.xlsx")
    ```

    **格式二：多工作表（dict 方式 — 推荐）**
    ```python
    data = {
        "员工信息": [["姓名", "年龄"], ["张三", 28]],
        "薪资表": [["姓名", "薪资"], ["张三", 10000]],
    }
    create_xlsx(data, "/tmp/公司数据.xlsx")
    ```

    **格式三：多工作表（list 方式 — 更灵活，支持逐 sheet 自定义样式）**
    ```python
    data = [
        {"sheet_name": "员工信息", "data": [...], "title": "2024年员工表"},
        {"sheet_name": "薪资表",   "data": [...],
         "header_fill": PatternFill(start_color="FFF2CC", fill_type="solid")},
    ]
    create_xlsx(data, "/tmp/公司数据.xlsx")
    ```

    **格式四：自定义全局样式**
    ```python
    from openpyxl.styles import Font, PatternFill
    data = [["姓名", "年龄"], ["张三", 28]]
    create_xlsx(data, "/tmp/人员.xlsx", styles={
        "header_fill": PatternFill(start_color="FFC000", fill_type="solid"),
        "even_fill": PatternFill(start_color="FFF8E1", fill_type="solid"),
    })
    ```

    Args:
        data: 数据内容，支持以下格式：
            - **list[list]**：单工作表，第一行为表头（向后兼容）
            - **dict[str, list[list]]**：多工作表，key 为 sheet 名
            - **list[dict]**：多工作表，每个 dict 包含：
                - sheet_name: str — 工作表名称
                - data: list[list] — 数据（第一行为表头）
                - title: str (可选) — 标题
                - header_font / header_fill / ... (可选) — 自定义样式
        output_path: 输出文件路径，以 .xlsx 结尾
        sheet_name: 单工作表模式下的 sheet 名称，默认 "Sheet1"
        title: 单工作表模式下的可选标题行
        auto_install: 是否自动安装 openpyxl（若未安装），默认 True
        styles: 全局样式配置（可选），支持以下键：
            - header_font: Font — 表头字体
            - header_fill: PatternFill — 表头背景色
            - header_alignment: Alignment — 表头对齐
            - data_font: Font — 数据字体
            - data_alignment: Alignment — 数据对齐
            - even_fill: PatternFill — 偶数行背景
            - odd_fill: PatternFill — 奇数行背景
            - border: Border — 单元格边框
            - title_font: Font — 标题字体
            - title_alignment: Alignment — 标题对齐

    Returns:
        成功或失败的消息

    Examples:
        >>> # 单工作表
        >>> create_xlsx(
        ...     data=[["姓名", "年龄"], ["张三", 28]],
        ...     output_path="/tmp/人员表.xlsx",
        ...     sheet_name="人员信息",
        ... )
        '✅ Excel 文件已生成: /tmp/人员表.xlsx (5.2 KB)'

        >>> # 多工作表
        >>> create_xlsx(
        ...     data={
        ...         "销售数据": [["产品", "销量"], ["苹果", 100]],
        ...         "库存数据": [["产品", "库存"], ["苹果", 500]],
        ...     },
        ...     output_path="/tmp/报表.xlsx",
        ... )
        '✅ Excel 文件已生成: /tmp/报表.xlsx (8.1 KB)'
    """
    # ── 第一步：自动安装 openpyxl（如需要） ──
    if auto_install:
        if not _auto_install_openpyxl():
            return "❌ Excel 生成失败: openpyxl 库未安装，自动安装也失败了"

    try:
        # ── 第二步：准备输出路径 ──
        output_path = _ensure_output_path(output_path)

        # ── 第三步：统一数据格式为 list[dict] 多 sheet 形式 ──
        sheet_defs: list[dict] = []

        if isinstance(data, dict):
            # dict 格式：{sheet_name: data_list}
            for s_name, s_data in data.items():
                if s_data and s_data[0]:
                    sheet_defs.append({
                        "sheet_name": s_name,
                        "data": s_data,
                    })
        elif isinstance(data, list):
            if not data:
                return "❌ Excel 生成失败: 数据不能为空"

            # 判断是单 sheet (list[list]) 还是多 sheet (list[dict])
            if data and isinstance(data[0], dict) and "data" in data[0]:
                # list[dict] 格式 — 多 sheet，每个元素包含详细配置
                for item in data:
                    sd = dict(item)
                    sd.setdefault("sheet_name", "Sheet1")
                    sheet_defs.append(sd)
            else:
                # list[list] 格式（向后兼容）
                if not data[0]:
                    return "❌ Excel 生成失败: 数据不能为空"
                sheet_defs.append({
                    "sheet_name": sheet_name,
                    "data": data,
                    "title": title,
                })
        else:
            return "❌ Excel 生成失败: 不支持的数据格式"

        # ── 第四步：创建工作簿并写入所有工作表 ──
        wb = Workbook()
        first_sheet = True

        for sheet_def in sheet_defs:
            data_rows = sheet_def.get("data", [])
            if not data_rows or not data_rows[0]:
                continue

            if first_sheet:
                # 使用现有的默认工作表并重命名
                ws = wb.active
                ws.title = sheet_def.get("sheet_name", "Sheet1")
                first_sheet = False
            else:
                ws = wb.create_sheet(
                    title=sheet_def.get("sheet_name", f"Sheet{len(wb.sheetnames) + 1}")
                )

            sheet_title = sheet_def.get("title")
            current_row = 1

            # 写入标题行
            if sheet_title:
                num_cols = len(data_rows[0])
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
                title_cell = ws.cell(row=1, column=1)
                title_cell.value = sheet_title

                tf = sheet_def.get("title_font") or (styles.get("title_font") if styles else None) or _TITLE_FONT
                ta = sheet_def.get("title_alignment") or (styles.get("title_alignment") if styles else None) or _TITLE_ALIGNMENT
                _apply_cell_style(title_cell, font=tf, alignment=ta)

                bdr = sheet_def.get("border") or (styles.get("border") if styles else None) or _THIN_BORDER
                for col_idx in range(1, num_cols + 1):
                    ws.cell(row=1, column=col_idx).border = bdr

                current_row = 2

            # 写入数据并应用格式（支持逐 sheet 自定义样式和全局样式）
            _apply_table_formatting(
                ws, data_rows, start_row=current_row,
                header_font=sheet_def.get("header_font") or (styles.get("header_font") if styles else None),
                header_fill=sheet_def.get("header_fill") or (styles.get("header_fill") if styles else None),
                header_alignment=sheet_def.get("header_alignment") or (styles.get("header_alignment") if styles else None),
                data_font=sheet_def.get("data_font") or (styles.get("data_font") if styles else None),
                data_alignment=sheet_def.get("data_alignment") or (styles.get("data_alignment") if styles else None),
                even_fill=sheet_def.get("even_fill") or (styles.get("even_fill") if styles else None),
                odd_fill=sheet_def.get("odd_fill") or (styles.get("odd_fill") if styles else None),
                border=sheet_def.get("border") or (styles.get("border") if styles else None),
            )

            # 自动调整列宽
            _auto_adjust_column_widths(ws, data_rows, title=sheet_title)

        # ── 第五步：保存文件 ──
        wb.save(output_path)

        size_str = _format_file_size(output_path)
        sheet_count = len(sheet_defs)
        if sheet_count == 1:
            return f"✅ Excel 文件已生成: {output_path} ({size_str})"
        else:
            sheet_names = ", ".join(s.get("sheet_name", "?") for s in sheet_defs)
            return f"✅ Excel 文件已生成: {output_path} ({size_str}) — 共 {sheet_count} 个工作表: [{sheet_names}]"

    except Exception as e:
        return f"❌ Excel 生成失败: {e}"

def _parse_md_table(md_table_content: str) -> tuple[list[str], list[list[str]]]:
    """解析 Markdown 表格文本，返回表头和数据行

    Args:
        md_table_content: Markdown 表格文本
            格式示例:
            | 姓名 | 年龄 | 城市 |
            | --- | --- | --- |
            | 张三 | 28 | 北京 |
            | 李四 | 35 | 上海 |

    Returns:
        (headers, rows) 元组
        - headers: 表头列表
        - rows: 数据行列表（每行是一个字符串列表）
    """
    lines = md_table_content.strip().split("\n")
    headers = []
    rows = []

    for i, line in enumerate(lines):
        stripped = line.strip()

        # 跳过空行
        if not stripped:
            continue

        # 跳过表格分隔行（如 | --- | --- |）
        if re.match(r"^\|[\s\-:|+]+\|$", stripped):
            continue

        # 确保是表格行（以 | 开头和结尾）
        if stripped.startswith("|") and stripped.endswith("|"):
            cells = [cell.strip() for cell in stripped.split("|")[1:-1]]

            if not headers:
                headers = cells
            else:
                rows.append(cells)

    return headers, rows


def convert_md_table_to_xlsx(
    md_table_content: str,
    output_path: str,
    sheet_name: str = "Sheet1",
    auto_install: bool = True,
) -> str:
    """将 Markdown 格式的表格文本转换为 .xlsx 文件

    自动解析 Markdown 表格语法，提取表头和数据行，
    并应用与 create_xlsx 相同的格式化样式（表头加粗、斑马纹、边框、自动列宽）。

    Args:
        md_table_content: Markdown 格式的表格内容字符串
            支持标准 GFM 表格格式:
            | 标题1 | 标题2 | 标题3 |
            | --- | --- | --- |
            | 数据1 | 数据2 | 数据3 |
            | 数据4 | 数据5 | 数据6 |
        output_path: 输出文件路径，需以 .xlsx 结尾
        sheet_name: 工作表名称，默认为 "Sheet1"
        auto_install: 是否自动安装 openpyxl，默认 True

    Returns:
        成功或失败的消息

    Examples:
        >>> md_table = \"\"\"| 产品 | 价格 | 库存 |
        ... | --- | --- | --- |
        ... | 苹果 | 5.0 | 100 |
        ... | 香蕉 | 3.5 | 200 |\"\"\"
        >>> convert_md_table_to_xlsx(md_table, "/tmp/产品表.xlsx")
        '✅ Excel 文件已生成: /tmp/产品表.xlsx (4.1 KB)'
    """
    try:
        # 自动安装（如需要）
        if auto_install:
            if not _auto_install_openpyxl():
                return "❌ Excel 生成失败: openpyxl 库未安装"

        # 解析 Markdown 表格
        headers, rows = _parse_md_table(md_table_content)

        if not headers:
            return "❌ Excel 生成失败: 未能解析到有效的 Markdown 表格内容"

        # 将解析结果组合为 create_xlsx 所需的数据格式
        data = [headers] + rows

        # 调用 create_xlsx 完成格式化和保存
        return create_xlsx(
            data=data,
            output_path=output_path,
            sheet_name=sheet_name,
            auto_install=False,  # 已安装过了
        )

    except Exception as e:
        return f"❌ Markdown 转 Excel 失败: {e}"
